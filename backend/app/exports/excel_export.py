import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.models import Timetable, ScheduledLesson, CurriculumEntry, TimeSlotConfig
from collections import defaultdict

DAYS_PT = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]


def hex_to_argb(hex_color: str) -> str:
    h = hex_color.lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    return "FF" + h.upper()


def generate_excel_timetable(db, timetable_id: int, view_type: str = "class", entity_id: int = None) -> bytes:
    tt = db.query(Timetable).filter(Timetable.id == timetable_id).first()
    if not tt:
        return b""

    q = db.query(ScheduledLesson).filter(ScheduledLesson.timetable_id == timetable_id)
    if view_type == "teacher" and entity_id:
        q = q.filter(ScheduledLesson.teacher_id == entity_id)
    elif view_type == "room" and entity_id:
        q = q.filter(ScheduledLesson.room_id == entity_id)
    elif view_type == "class" and entity_id:
        q = q.join(CurriculumEntry, ScheduledLesson.curriculum_entry_id == CurriculumEntry.id)
        q = q.filter(CurriculumEntry.class_id == entity_id)

    lessons = q.all()

    slot_rows = db.query(TimeSlotConfig).filter(
        TimeSlotConfig.academic_year_id == tt.academic_year_id,
        TimeSlotConfig.is_break == False  # noqa
    ).all()
    all_slots = sorted({r.slot_number for r in slot_rows})
    slot_times = {}
    for r in slot_rows:
        if r.slot_number not in slot_times:
            slot_times[r.slot_number] = f"{r.start_time.strftime('%H:%M')}-{r.end_time.strftime('%H:%M')}"

    grid = defaultdict(dict)
    for lesson in lessons:
        entry = lesson.curriculum_entry
        subject_name = entry.subject.name if entry and entry.subject else "?"
        subject_color = entry.subject.color if entry and entry.subject else "#3498db"
        teacher_name = lesson.teacher.name if lesson.teacher else ""
        class_name = entry.class_.name if entry and entry.class_ else ""

        if view_type == "teacher":
            label = f"{subject_name}\n{class_name}"
        elif view_type == "room":
            label = f"{subject_name}\n{class_name} / {teacher_name}"
        else:
            label = f"{subject_name}\n{teacher_name}"

        grid[lesson.slot_number][lesson.day_of_week] = (label, subject_color)

    wb = Workbook()
    ws = wb.active
    ws.title = "Horário"

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Header row
    ws.cell(row=1, column=1, value="Hora").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).border = thin

    for col, day_name in enumerate(DAYS_PT, start=2):
        cell = ws.cell(row=1, column=col, value=day_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.column_dimensions["A"].width = 12
    ws.row_dimensions[1].height = 20

    for row_idx, slot in enumerate(all_slots, start=2):
        time_str = slot_times.get(slot, str(slot))
        time_cell = ws.cell(row=row_idx, column=1, value=time_str)
        time_cell.alignment = center
        time_cell.border = thin
        time_cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        ws.row_dimensions[row_idx].height = 40

        for day in range(5):
            cell = ws.cell(row=row_idx, column=day + 2)
            cell.border = thin
            cell.alignment = center
            if slot in grid and day in grid[slot]:
                label, color = grid[slot][day]
                cell.value = label
                argb = hex_to_argb(color)
                cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
